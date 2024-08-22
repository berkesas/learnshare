import openpyxl
import json

# Give the location of the file
path = "chatbot_intents.xlsx"

# To open the workbook
# workbook object is created
wb_obj = openpyxl.load_workbook(path)

# Get workbook active sheet object
# from the active attribute
sheet_obj = wb_obj.active

row = sheet_obj.max_row
column = sheet_obj.max_column

print("Total Rows:", row)
print("Total Columns:", column)

# create an intents list
intents = {}
# Loop over intents to create a list
# of intents start row is 2 because there is a
# header row
for i in range(2, row + 1):
    cell_obj = sheet_obj.cell(row=i, column=1)
    if cell_obj.value and len(cell_obj.value) > 0:
        if cell_obj.value not in intents:
            print('adding', cell_obj.value)
            intents[cell_obj.value] = {}
            intents[cell_obj.value]["tag"] = cell_obj.value
            intents[cell_obj.value]["proposable"] = sheet_obj.cell(row=i, column=4).value
            intents[cell_obj.value]["patterns"] = []
            intents[cell_obj.value]["responses"] = []
    else:
        break

    # add questions and responses
    intents[cell_obj.value]["patterns"].append(
        sheet_obj.cell(row=i, column=2).value)
    intents[cell_obj.value]["responses"].append(
        sheet_obj.cell(row=i, column=3).value)

# create a JSON object for our data
data = {
    "intents": list(intents.values())
}

# format data
json_data = json.dumps(data, indent=4)
# print(json_data)
# print(intents.values())

# write json data to file
with open('chatbot_intents.json', 'w') as json_file:
    json_file.write(json_data)
